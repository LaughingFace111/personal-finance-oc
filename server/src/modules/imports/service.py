import csv
import io
import json
from datetime import datetime
from decimal import Decimal
from typing import List, Optional, Dict, Any
from dateutil import parser as date_parser
from sqlalchemy.exc import IntegrityError  # 🛡️ L: 添加唯一约束异常捕获

from sqlalchemy.orm import Session

from src.common.enums import ImportStatus, ConfirmStatus, TransactionType, TransactionDirection, SourceType
from src.core import generate_uuid, NotFoundException
from src.common import safe_decimal, safe_int

from .models import ImportBatch, ImportRow
from .schemas import UpdateImportRowRequest, ConfirmImportRequest
from src.modules.transactions.service import create_transaction
from src.modules.transactions.schemas import TransactionCreate
from src.modules.accounts.service import get_accounts
from src.modules.bills.parsers.wechat import WechatBillParser
from src.modules.categories.service import get_categories


# Standard import fields
STANDARD_FIELDS = [
    "occurred_at", "posted_at", "amount", "direction", "account_name",
    "merchant", "description", "counterparty", "category", "external_txn_id"
]


def parse_wechat_csv(content: bytes):
    """兼容旧导入链路，复用账单模块中的微信 CSV 解析器。"""
    return WechatBillParser().parse_csv(content)


def _parse_date(date_str: str) -> datetime:
    """Parse date string to datetime"""
    if not date_str:
        return datetime.utcnow()
    try:
        return date_parser.parse(date_str)
    except:
        return datetime.utcnow()


def _guess_transaction_type(row: Dict, categories: List) -> tuple:
    """Guess transaction type based on row data"""
    amount = safe_decimal(row.get("amount", 0))
    direction = row.get("direction", "").lower()
    description = row.get("description", "").lower()
    merchant = row.get("merchant", "").lower()

    # Determine direction and type
    if direction in ["in", "收入"]:
        return TransactionType.INCOME, TransactionDirection.IN
    elif direction in ["out", "支出"]:
        return TransactionType.EXPENSE, TransactionDirection.OUT
    elif direction in ["transfer", "转账"]:
        return TransactionType.TRANSFER, TransactionDirection.INTERNAL

    # Auto-guess based on keywords
    income_keywords = ["工资", "收入", "奖金", "退款", "报销"]
    expense_keywords = ["消费", "支出", "购物", "餐饮", "交通"]

    text = description + merchant
    if any(k in text for k in income_keywords):
        return TransactionType.INCOME, TransactionDirection.IN

    if amount and amount < 0:
        return TransactionType.EXPENSE, TransactionDirection.OUT

    return TransactionType.EXPENSE, TransactionDirection.OUT


def _guess_category(row: Dict, categories: List) -> Optional[str]:
    """Guess category based on merchant/description"""
    description = row.get("description", "").lower()
    merchant = row.get("merchant", "").lower()
    text = description + merchant

    # Simple keyword matching
    keywords_map = {
        "餐饮": ["餐饮", "吃饭", "外卖", "餐厅"],
        "交通": ["交通", "地铁", "公交", "打车", "油费"],
        "购物": ["购物", "淘宝", "京东", "天猫"],
        "工资": ["工资", "薪资"],
    }

    for category_name, keywords in keywords_map.items():
        if any(k in text for k in keywords):
            for cat in categories:
                if cat.name == category_name:
                    return cat.id

    return None


def _parse_csv(content: str) -> List[Dict]:
    """Parse CSV content to list of dicts"""
    lines = content.strip().split("\n")
    if not lines:
        return []

    reader = csv.DictReader(io.StringIO("\n".join(lines)))
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> List[Dict]:
    """Parse XLSX content to list of dicts"""
    try:
        from openpyxl import load_workbook
        workbook = io.BytesIO(content)
        wb = load_workbook(workbook, data_only=True)
        sheet = wb.active
        
        # Get headers from first row
        headers = []
        for cell in sheet[1]:
            headers.append(cell.value)
        
        # Read data rows
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(row):  # Skip empty rows
                continue
            row_dict = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_dict[headers[i]] = str(value) if value is not None else ""
            rows.append(row_dict)
        
        return rows
    except Exception as e:
        print(f"XLSX parse error: {e}")
        return []


def create_import_batch(db: Session, book_id: str, filename: str, content: bytes, source_name: str = None) -> ImportBatch:
    """Create import batch and parse file"""
    file_type = "csv" if filename.endswith(".csv") else "xlsx"

    # Create batch
    batch = ImportBatch(
        id=generate_uuid(),
        book_id=book_id,
        filename=filename,
        source_name=source_name,
        file_type=file_type,
        status=ImportStatus.UPLOADED.value,
    )
    db.add(batch)

    # Parse content based on file type
    if file_type == "xlsx":
        rows = _parse_xlsx(content)
    else:
        # Try to decode as UTF-8 for CSV
        try:
            text = content.decode("utf-8")
            rows = _parse_csv(text)
        except UnicodeDecodeError:
            rows = []
    
    batch.total_rows = len(rows)

    # Get accounts and categories for guessing
    accounts = get_accounts(db, book_id)
    categories = get_categories(db, book_id)

    # Create import rows
    import_rows = []
    for idx, row in enumerate(rows):
        # Normalize data
        occurred_at = _parse_date(row.get("occurred_at", ""))
        amount = safe_decimal(row.get("amount", 0))
        if amount and amount < 0:
            amount = abs(amount)

        # Guess transaction type
        tx_type, direction = _guess_transaction_type(row, categories)

        # Guess category
        guessed_category_id = _guess_category(row, categories)

        # Guess account
        guessed_account_id = None
        account_name = row.get("account_name", "")
        for acc in accounts:
            if account_name and account_name.lower() in acc.name.lower():
                guessed_account_id = acc.id
                break

        # Create normalized data
        normalized = {
            "occurred_at": occurred_at.isoformat() if occurred_at else None,
            "amount": str(amount) if amount else None,
            "direction": direction.value if hasattr(direction, 'value') else direction,
            "transaction_type": tx_type.value if hasattr(tx_type, 'value') else tx_type,
            "merchant": row.get("merchant", ""),
            "description": row.get("description", ""),
            "counterparty": row.get("counterparty", ""),
            "category": row.get("category", ""),
        }

        import_row = ImportRow(
            id=generate_uuid(),
            batch_id=batch.id,
            row_no=idx + 1,
            raw_data=json.dumps(row, ensure_ascii=False),
            normalized_data=json.dumps(normalized, ensure_ascii=False),
            guessed_account_id=guessed_account_id,
            guessed_category_id=guessed_category_id,
            guessed_transaction_type=tx_type.value if hasattr(tx_type, 'value') else tx_type,
            guessed_confidence=Decimal("70") if guessed_category_id else Decimal("30"),
            confirm_status=ConfirmStatus.PENDING.value,
        )
        import_rows.append(import_row)

    db.add_all(import_rows)
    batch.parsed_rows = len(import_rows)
    batch.status = ImportStatus.PARSED.value

    db.commit()
    db.refresh(batch)
    return batch


def get_import_batches(db: Session, book_id: str) -> List[ImportBatch]:
    """Get import batches"""
    return db.query(ImportBatch).filter(
        ImportBatch.book_id == book_id
    ).order_by(ImportBatch.created_at.desc()).all()


def get_import_batch(db: Session, batch_id: str, book_id: str) -> Optional[ImportBatch]:
    """Get import batch by ID"""
    return db.query(ImportBatch).filter(
        ImportBatch.id == batch_id,
        ImportBatch.book_id == book_id
    ).first()


def get_import_rows(db: Session, batch_id: str, confirm_status: str = None) -> List[ImportRow]:
    """Get import rows"""
    query = db.query(ImportRow).filter(ImportRow.batch_id == batch_id)
    if confirm_status:
        query = query.filter(ImportRow.confirm_status == confirm_status)
    return query.order_by(ImportRow.row_no).all()


def update_import_row(db: Session, row_id: str, data: UpdateImportRowRequest) -> ImportRow:
    """Update import row"""
    row = db.query(ImportRow).filter(ImportRow.id == row_id).first()
    if not row:
        raise NotFoundException("Import row not found")

    if data.guessed_account_id:
        row.guessed_account_id = data.guessed_account_id
    if data.guessed_category_id:
        row.guessed_category_id = data.guessed_category_id
    if data.guessed_transaction_type:
        row.guessed_transaction_type = data.guessed_transaction_type
    if data.confirm_status:
        row.confirm_status = data.confirm_status

    row.user_modified = "true"

    db.commit()
    db.refresh(row)
    return row


def confirm_import(db: Session, batch_id: str, book_id: str, data: ConfirmImportRequest) -> Dict:
    """Confirm import and create transactions"""
    batch = get_import_batch(db, batch_id, book_id)
    if not batch:
        raise NotFoundException("Import batch not found")

    # Get all rows for this batch first
    all_rows = db.query(ImportRow).filter(ImportRow.batch_id == batch_id).all()
    
    # Get rows to import
    if data.confirmed_row_ids:
        rows_to_process = [r for r in all_rows if r.id in data.confirmed_row_ids]
    else:
        rows_to_process = [r for r in all_rows if r.confirm_status == ConfirmStatus.PENDING.value]

    # Get accounts for lookup
    accounts = {acc.id: acc for acc in get_accounts(db, book_id)}
    categories = {cat.id: cat for cat in get_categories(db, book_id)}

    # Initialize counters
    confirmed_count = 0
    skipped_count = 0
    duplicate_count = 0
    error_count = 0

    # Use explicit transaction for atomicity
    try:
        with db.begin():
            for row in rows_to_process:
                # 🛡️ L: 嵌套事务 savepoint — 单行失败仅回滚当前层，不污染整批
                try:
                    with db.begin_nested():
                        if not row.normalized_data:
                            row.confirm_status = ConfirmStatus.SKIPPED.value
                            skipped_count += 1
                            continue

                        normalized = json.loads(row.normalized_data)
                        if not normalized.get("amount"):
                            row.confirm_status = ConfirmStatus.SKIPPED.value
                            skipped_count += 1
                            continue

                        account = accounts.get(row.guessed_account_id)
                        if not account:
                            row.error_message = "Account not found"
                            row.confirm_status = ConfirmStatus.SKIPPED.value
                            error_count += 1
                            continue

                        if normalized.get("external_txn_id"):
                            business_key = f"external:{row.guessed_account_id}:{normalized['external_txn_id']}"
                        else:
                            import hashlib
                            raw_hash = hashlib.sha256(row.raw_data.encode()).hexdigest()[:16]
                            business_key = f"import:{book_id}:{raw_hash}"

                        tx_type = normalized.get("transaction_type", "expense")
                        direction = normalized.get("direction", "out")

                        txn_data = TransactionCreate(
                            occurred_at=datetime.fromisoformat(normalized["occurred_at"]) if normalized.get("occurred_at") else datetime.utcnow(),
                            transaction_type=tx_type,
                            direction=direction,
                            amount=Decimal(normalized["amount"]),
                            account_id=row.guessed_account_id,
                            category_id=row.guessed_category_id,
                            merchant=normalized.get("merchant"),
                            note=normalized.get("description"),
                            source_type=SourceType.IMPORT,
                            source_batch_id=batch_id,
                            source_row_no=row.row_no,
                            business_key=business_key,
                            include_expense_override=(tx_type == "expense" or direction == "out"),
                            include_income_override=(tx_type == "income" or direction == "in"),
                            include_cashflow_override=True,
                        )

                        create_transaction(db, book_id, txn_data)
                        row.confirm_status = ConfirmStatus.CONFIRMED.value
                        confirmed_count += 1

                except IntegrityError:
                    # 唯一键冲突 → 仅标记重复，继续下一行
                    row.error_message = "重复数据"
                    row.confirm_status = ConfirmStatus.DUPLICATE.value
                    duplicate_count += 1
                    continue

                except Exception as e:
                    row.error_message = str(e)
                    row.confirm_status = ConfirmStatus.SKIPPED.value
                    error_count += 1
                    skipped_count += 1

            # Update batch stats
            batch.confirmed_rows = confirmed_count
            batch.skipped_rows = skipped_count
            batch.duplicate_rows = duplicate_count
            batch.status = ImportStatus.CONFIRMED.value if confirmed_count > 0 else ImportStatus.FAILED.value

    except Exception as e:
        db.rollback()
        raise AppException(status_code=500, code=50001, message=f"Import failed: {str(e)}")

    return {
        "batch": batch,
        "confirmed_count": confirmed_count,
        "skipped_count": skipped_count,
        "duplicate_count": duplicate_count,
        "error_count": error_count
    }
