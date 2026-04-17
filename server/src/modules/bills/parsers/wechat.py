import csv
import io
import re
import zipfile
from datetime import datetime
from decimal import Decimal
from typing import Dict, List, Sequence

from src.common.enums import TransactionDirection, TransactionType

from .base import BillParser, BillRecord


class WechatBillParser(BillParser):
    """微信支付账单解析器"""

    KEY_HEADER_COLS = {
        "交易时间",
        "收/支",
        "金额(元)",
    }
    REQUIRED_HEADERS = {
        "交易时间",
        "交易类型",
        "交易对方",
        "商品",
        "收/支",
        "金额(元)",
        "当前状态",
        "交易单号",
    }
    ENCODINGS = ("utf-8-sig", "gbk", "gb2312")

    ACCEPTED_STATUS = {
        "支付成功",
        "已收款",
        "已存入零钱",
        "转账成功",
        "收款成功",
        "对方已收钱",
        "对方已收款",
    }

    IGNORE_KEYWORDS: List[str] = []

    INCOME_KEYWORDS = [
        "转账收入",
        "收款",
        "红包收入",
        "群收款",
    ]

    EXPENSE_KEYWORDS = [
        "消费",
        "支付",
        "扫码支付",
        "付款",
    ]

    def parse(self, content: bytes) -> List[BillRecord]:
        try:
            if self._looks_like_xlsx(content):
                return self._parse_xlsx(content)
            raise ValueError("微信账单仅支持 xlsx 格式，请从微信导出 xlsx 格式账单")
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError(f"微信账单解析失败: {exc}") from exc

    def _parse_xlsx(self, content: bytes) -> List[BillRecord]:
        try:
            if not zipfile.is_zipfile(io.BytesIO(content)):
                raise ValueError("文件不是有效的 xlsx 格式")

            try:
                import openpyxl
            except ImportError as exc:
                raise ValueError("当前环境未安装 openpyxl") from exc

            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            ws = wb.active
            header_index = self._find_sheet_header_index(ws)
            if header_index < 0:
                raise ValueError("未找到微信账单表头，请确认文件格式")

            headers = self._extract_headers(
                [self._normalize_cell_value(cell.value) for cell in ws[header_index]]
            )
            rows: List[Dict[str, str]] = []
            for row_idx in range(header_index + 1, ws.max_row + 1):
                row_data: Dict[str, str] = {}
                for col_idx, header in enumerate(headers, start=1):
                    if not header:
                        continue
                    row_data[header] = self._normalize_cell_value(ws.cell(row_idx, col_idx).value)
                rows.append(row_data)

            return self._parse_rows(rows)
        except ValueError:
            raise
        except Exception as exc:
            raise ValueError(f"微信账单 XLSX 解析失败: {exc}") from exc

    def _parse_rows(self, rows: Sequence[Dict[str, str]]) -> List[BillRecord]:
        raw_records: List[BillRecord] = []
        row_num = 1

        for row_data in rows:
            if not row_data or not row_data.get("交易时间"):
                continue

            status = row_data.get("当前状态", "")
            if status not in self.ACCEPTED_STATUS:
                continue

            trade_type = row_data.get("交易类型", "")
            ignore_reason = self._check_ignore(trade_type, row_data)
            if ignore_reason:
                continue

            try:
                occurred_at = self._parse_datetime(row_data.get("交易时间", ""))
                amount = self._parse_amount(row_data.get("金额(元)", ""))
                direction_raw = self._normalize_cell_value(row_data.get("收/支", ""))
                # 当 direction_raw 为 "/"、"不计收支" 或空时，传入空字符串给 _resolve_direction 通过关键词判断
                in_out_for_resolve = "" if direction_raw in {"/", "不计收支", ""} else direction_raw
                direction, tx_type = self._resolve_direction(
                    in_out_for_resolve,
                    trade_type,
                    row_data.get("商品", ""),
                )

                raw_records.append(
                    BillRecord(
                        row_no=row_num,
                        occurred_at=occurred_at,
                        transaction_type=tx_type,
                        direction=direction,
                        amount=amount,
                        counterparty=row_data.get("交易对方") or None,
                        counterparty_account=None,
                        description=row_data.get("商品") or None,
                        category=None,
                        in_out=row_data.get("收/支", ""),
                        status=status,
                        transaction_order_no=row_data.get("交易单号", ""),
                        merchant_order_no=row_data.get("商户单号", "") or None,
                        payment_method=row_data.get("支付方式") or None,
                        note=row_data.get("备注") or None,
                    )
                )
                row_num += 1
            except ValueError:
                continue

        if not raw_records:
            raise ValueError("未解析到有效的微信账单记录，请确认文件内容与账单类型")

        return raw_records

    def _find_sheet_header_index(self, ws) -> int:
        for row_idx in range(1, min(25, ws.max_row + 1)):
            row_values = [
                self._normalize_cell_value(ws.cell(row_idx, col).value)
                for col in range(1, ws.max_column + 1)
            ]
            if self._is_header_row(row_values):
                return row_idx
        return -1

    def _is_header_row(self, row_values: Sequence[str]) -> bool:
        normalized = {value for value in row_values if value}
        return self.KEY_HEADER_COLS.issubset(normalized)

    def _is_header_line(self, line: str) -> bool:
        try:
            values = [self._normalize_cell_value(value) for value in next(csv.reader([line]))]
        except (csv.Error, StopIteration):
            return False
        present = {value for value in values if value}
        return "交易时间" in present and "金额(元)" in present

    def _extract_headers(self, row_values: Sequence[str]) -> List[str]:
        headers = [self._normalize_cell_value(value) for value in row_values]
        normalized = {value for value in headers if value}
        if not self.REQUIRED_HEADERS.issubset(normalized):
            missing_headers = sorted(self.REQUIRED_HEADERS - normalized)
            raise ValueError(f"微信账单表头缺少必要列: {', '.join(missing_headers)}")
        return headers

    def _looks_like_xlsx(self, content: bytes) -> bool:
        return content[:2] == b"PK"

    def _normalize_cell_value(self, value) -> str:
        if value is None:
            return ""
        return str(value).replace("\ufeff", "").strip()

    def _parse_datetime(self, value: str) -> datetime:
        if not value:
            raise ValueError("交易时间为空")
        value = str(value).strip()
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M:%S.0", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d"]:
            try:
                dt = datetime.strptime(value, fmt)
                return datetime(dt.year, dt.month, dt.day)
            except ValueError:
                continue
        raise ValueError(f"无法解析日期: {value}")

    def _parse_amount(self, value) -> Decimal:
        if value is None:
            raise ValueError("金额为空")

        normalized = (
            str(value)
            .replace("¥", "")
            .replace("￥", "")
            .replace("元", "")
            .replace(",", "")
            .replace(" ", "")
            .replace("\xa0", "")
            .replace("\u3000", "")
            .strip()
        )
        if not normalized:
            raise ValueError("金额为空")

        normalized = re.sub(r"[^\d.\-+]", "", normalized)
        if normalized in {"", "+", "-", ".", "+.", "-."}:
            raise ValueError("金额格式无效")

        amount = Decimal(normalized)
        return abs(amount)

    def _resolve_direction(self, in_out: str, trade_type: str, description: str) -> tuple:
        in_out = in_out.strip() if in_out else ""

        if in_out == "收入":
            return TransactionDirection.IN, TransactionType.INCOME
        if in_out == "支出":
            return TransactionDirection.OUT, TransactionType.EXPENSE

        # in_out 为空（来自 "/"、"不计收支" 或空值），通过 trade_type 关键词推断方向
        if not in_out:
            text = f"{trade_type} {description}".lower()

            # 首先检查退款
            if "退款" in trade_type or "退款" in description:
                return TransactionDirection.IN, TransactionType.INCOME

            # 通过 trade_type 推断收入类
            for keyword in ["转入", "收入", "收款", "红包", "转账"]:
                if keyword in trade_type:
                    return TransactionDirection.IN, TransactionType.INCOME

            # 通过 trade_type 推断支出类
            for keyword in ["转出", "消费", "支付", "扫码", "付款", "充值", "提现"]:
                if keyword in trade_type:
                    return TransactionDirection.OUT, TransactionType.EXPENSE

            raise ValueError(
                f"无法判断收支方向: in_out={in_out or '<empty>'}, trade_type={trade_type}, description={description}"
            )

        text = f"{trade_type} {description}".lower()

        for keyword in self.INCOME_KEYWORDS:
            if keyword in text:
                return TransactionDirection.IN, TransactionType.INCOME

        for keyword in self.EXPENSE_KEYWORDS:
            if keyword in text:
                return TransactionDirection.OUT, TransactionType.EXPENSE

        if "退款" in trade_type or "退款" in description:
            return TransactionDirection.IN, TransactionType.INCOME

        raise ValueError(
            f"无法判断收支方向: in_out={in_out or '<empty>'}, trade_type={trade_type}, description={description}"
        )

    def _check_ignore(self, trade_type: str, row_data: Dict) -> str:
        # 所有交易都要计入，永远不跳过
        return ""
